from django.http import HttpResponse
from django.db.models.query import QuerySet
from openpyxl import Workbook
from StringIO import StringIO
import csv


class SpreadsheetResponseMixin(object):
    filename_base = 'export'

    def render_excel_response(self, **kwargs):
        filename = self.get_filename(extension='xlsx')
        # Generate content
        self.data, self.headers = self.render_setup(**kwargs)
        # Setup response
        content_type = \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        # Add content and return response
        self.generate_xlsx(data=self.data, headers=self.headers, file=response)
        return response

    def render_csv_response(self, **kwargs):
        filename = self.get_filename(extension='csv')
        # Generate content
        self.data, self.headers = self.render_setup(**kwargs)
        # Build response
        content_type = 'text/csv'
        response = HttpResponse(content_type=content_type)
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)
        # Add content to response
        self.generate_csv(data=self.data, headers=self.headers, file=response)
        return response

    def render_setup(self, **kwargs):
        # Generate content
        if 'queryset' in kwargs:
            self.queryset = kwargs['queryset']

        if not hasattr(self, 'queryset') and 'model' in kwargs:
            self.queryset = model.objects.all()

        if not hasattr(self, 'queryset') and hasattr(self, 'model'):
            self.queryset = self.model.objects.all()

        if not hasattr(self, 'queryset'):
            raise NotImplementedError(
                "You must provide a queryset or model on the class, or pass one in."
            )

        fields = self.get_fields(**kwargs)
        data = self.generate_data(fields=fields)

        headers = kwargs.get('headers')
        if not headers:
            headers = self.generate_headers(self.queryset.model, fields=fields)

        return data, headers

    def recursively_extract_value(self, current_instance, remaining_path):
        if '__' in remaining_path:
            foreign_key_name, path_in_related_instance = remaining_path.split('__', 2)
            related_instance = getattr(current_instance, foreign_key_name)
            return self.recursively_extract_value(related_instance, path_in_related_instance)
        else:
            return getattr(current_instance, remaining_path)

    def generate_data(self, fields=None):
        # After all that, have we got a proper queryset?
        assert isinstance(self.queryset, QuerySet)

        if getattr(self, 'use_models', False):
            fields = self.get_fields(fields=fields)
            return self.generate_data_using_models(fields)
        elif fields:
            return self.generate_data_using_fields(fields)
        else:
            return self.generate_data_using_values()

    def generate_data_using_models(self, fields):
        for model_instance in self.queryset:
            row = []

            for field in fields:
                calculated = self.get_calculated_field(field)
                if calculated:
                    value = calculated(model_instance)
                else:
                    value = self.recursively_extract_value(model_instance, field)
                    if value and callable(value):
                        value = value()
                row.append(value)

            yield tuple(row)

    def generate_data_using_fields(self, fields):
        columns = []

        # For each field, contains the virtual field name, and the starting
        # offset and length of the database columns used to evaluate it.
        # If the field is calculated (a method on self), it can have any
        # length, otherwise the length will be 1, and the value returned by
        # values_list() will be indexed at that location and returned directly.
        field_maps = []

        for field in fields:
            calculated = self.get_calculated_field(field)

            if calculated:
                field_map = (field, calculated, len(columns))
                columns += calculated.fields
                field_maps.append(field_map)
            else:
                field_map = (field, None, len(columns))
                columns.append(field)
                field_maps.append(field_map)

        for row in self.queryset.values_list(*columns):
            values_out = []
            for field, calculated, offset in field_maps:
                if calculated is None:
                    values_out.append(row[offset])
                else:
                    length = len(calculated.fields)
                    values_out.append(calculated(row[offset:offset+length]))
            yield tuple(values_out)

    def generate_data_using_values(self):
        for row in self.queryset.values_list():
            yield row

    def recursively_build_field_name(self, current_model, remaining_path):
        get_field = lambda name: current_model._meta.get_field(name)

        if '__' in remaining_path:
            foreign_key_name, path_in_related_model = remaining_path.split('__', 2)
            foreign_key_field = get_field(foreign_key_name)
            related_model = foreign_key_field.rel.to
            return [unicode(foreign_key_field.verbose_name)] + \
                self.recursively_build_field_name(related_model, path_in_related_model)
        else:
            return [unicode(get_field(remaining_path).verbose_name)]

    def get_calculated_field(self, field_name):
        calculated_field = getattr(self, field_name, None)
        if calculated_field and callable(calculated_field):
            return calculated_field
        else:
            return None

    def build_field_name(self, model, path):
        calculated_field = self.get_calculated_field(path)
        if calculated_field:
            if hasattr(calculated_field, 'verbose_name'):
                return calculated_field.verbose_name
            else:
                return path.replace('_', ' ').title()
        else:
            name_parts = self.recursively_build_field_name(model, path)
            return ' '.join(name_parts).title()

    def generate_headers(self, model, fields):
        return tuple(self.build_field_name(model, field) for field in fields)

    def generate_xlsx(self, data, headers=None, file=None):
        wb = Workbook()
        ws = wb.get_active_sheet()

        # Put in headers
        rowoffset = 0
        if headers:
            rowoffset = 1
            for c, headerval in enumerate(headers, 1):
                ws.cell(row=1, column=c).value = headerval

        # Put in data
        for r, row in enumerate(data, 1):
            for c, cellval in enumerate(row, 1):
                ws.cell(row=r + rowoffset, column=c).value = cellval
        if file:
            wb.save(file)
        return wb

    def generate_csv(self, data, headers=None, file=None):
        if not file:
            generated_csv = StringIO()
        else:
            generated_csv = file
        writer = csv.writer(generated_csv, dialect='excel')
        # Put in headers
        if headers:
            writer.writerow([unicode(s).encode('utf-8') for s in headers])

        # Put in data
        for row in data:
            writer.writerow([unicode(s).encode('utf-8') for s in row])
        return generated_csv

    def get_render_method(self, format):
        if format == 'excel':
            return self.render_excel_response
        elif format == 'csv':
            return self.render_csv_response
        raise NotImplementedError("Export format is not recognized.")

    def get_format(self, **kwargs):
        if 'format' in kwargs:
            return kwargs['format']
        elif hasattr(self, 'format'):
            return self.format
        raise NotImplementedError("Format is not defined.")

    def get_filename(self, **kwargs):
        if 'filename' in kwargs:
            return kwargs['filename']
        if hasattr(self, 'filename'):
            return self.filename

        extension = kwargs.get('extension', 'out')
        return "{0}.{1}".format(self.filename_base, extension)

    def get_fields(self, model=None, **kwargs):
        if 'fields' in kwargs:
            return kwargs['fields']
        elif hasattr(self, 'fields') and self.fields is not None:
            return self.fields
        elif hasattr(self.queryset, 'field_names'):
            return self.queryset.field_names
        else:
            model = self.queryset.model
            return [f.name for f in model._meta.fields]
