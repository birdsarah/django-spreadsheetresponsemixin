# -*- coding: utf-8 -*-
from django.http import HttpResponse
from django.test import TestCase
from StringIO import StringIO
import mock
import pytest
import factory
from openpyxl import Workbook

from spreadsheetresponsemixin import SpreadsheetResponseMixin
from .models import MockModel, MockAuthor


class MockModelFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = MockModel
    title = factory.Sequence(lambda n: 'title{0}'.format(n))


class MockAuthorFactory(factory.django.DjangoModelFactory):
    class Meta:
        model = MockAuthor
    name = factory.Sequence(lambda n: 'name{0}'.format(n))


class GenerateDataTests(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()
        self.author = MockAuthorFactory()
        self.mock = MockModelFactory(author=self.author)
        self.mock2 = MockModelFactory(author=self.author)
        self.queryset = MockModel.objects.all()

    def test_assertion_error_raised_if_not_a_queryset_is_sent(self):
        with pytest.raises(AssertionError):
            self.mixin.queryset = [1]
            list(self.mixin.generate_data())

    def test_if_queryset_is_none_gets_self_queryset(self):
        self.mixin.queryset = MockModel.objects.all()
        self.assertSequenceEqual(MockModel.objects.values_list(),
                                 list(self.mixin.generate_data()))

    def test_returns_values_list_qs_if_queryset(self):
        self.mixin.queryset = self.queryset
        expected_list = self.queryset.values_list()
        actual_list = self.mixin.generate_data()
        assert list(actual_list) == list(expected_list)

    def test_returns_values_list_if_qs_is_values_queryset(self):
        self.mixin.queryset = self.queryset.values()
        expected_list = MockModel.objects.all().values_list()
        actual_list = self.mixin.generate_data()
        assert list(actual_list) == list(expected_list)

    def test_returns_self_if_qs_is_values_list_queryset(self):
        values_list_queryset = MockModel.objects.all().values_list()
        self.mixin.queryset = values_list_queryset
        actual_list = self.mixin.generate_data()
        assert list(actual_list) == list(values_list_queryset)

    def test_uses_specified_fields(self):
        fields = ('title',)
        values_list_queryset = MockModel.objects.all().values_list()
        self.mixin.queryset = values_list_queryset
        # We expect it to be filtered by title,
        # even though full value_list is passed
        expected_list = MockModel.objects.all().values_list(*fields)
        actual_list = self.mixin.generate_data(fields)
        assert list(actual_list) == list(expected_list)

    def test_follows_foreign_key_with_model_queryset(self):
        fields = ('title', 'author__name')
        queryset = MockModel.objects.all()
        self.mixin.queryset = queryset
        expected_list = [
            (self.mock.title, self.author.name),
            (self.mock2.title, self.author.name),
        ]
        actual_list = self.mixin.generate_data(fields)
        assert list(actual_list) == list(expected_list)

    def test_allows_calculated_field_values(self):
        fields = ('title', 'author__name', 'calculated')
        queryset = MockModel.objects.all()
        self.mixin.queryset = queryset
        expected_list = [
            (self.mock.title, self.author.name, u'whee %d' % self.mock.id),
            (self.mock2.title, self.author.name, u'whee %d' % self.mock2.id),
        ]

        self.mixin.calculated = lambda values: 'whee %d' % values[0]
        self.mixin.calculated.fields = ['id']

        actual_list = self.mixin.generate_data(fields)
        self.assertEqual(list(actual_list), expected_list)

    def test_follows_foreign_key_with_values_list_queryset(self):
        fields = ('title', 'author__name')
        values_list_queryset = MockModel.objects.all().values_list()
        self.mixin.queryset = values_list_queryset
        expected_list = [
            (self.mock.title, self.author.name),
            (self.mock2.title, self.author.name),
        ]
        actual_list = self.mixin.generate_data(fields)
        assert list(actual_list) == list(expected_list)

    def test_reverse_ordering_when_fields_specified(self):
        fields = ('title', 'id')
        self.mixin.queryset = self.queryset
        actual_list = self.mixin.generate_data(fields)
        assert list(actual_list)[0] == (self.mock.title, self.mock.id)

    def test_allows_evaluation_using_models(self):
        fields = ('title', 'author__name', 'calculated')
        self.mixin.queryset = self.queryset
        expected_list = [
            (self.mock.title, self.author.name, u'whee %d' % self.mock.id),
            (self.mock2.title, self.author.name, u'whee %d' % self.mock2.id),
        ]

        self.mixin.use_models = True
        self.mixin.calculated = lambda model: 'whee %d' % model.id

        actual_list = self.mixin.generate_data(fields)
        self.assertEqual(list(actual_list), expected_list)


class GenerateXlsxTests(TestCase):
    def setUp(self):
        self.data = (('row1col1', 'row1col2'), ('row2col1', 'row2col2'))
        self.mixin = SpreadsheetResponseMixin()

    def _get_sheet(self, wb):
        return wb.get_active_sheet()

    def test_returns_workbook_if_no_file_passed(self):
        assert type(self.mixin.generate_xlsx(self.data)) == Workbook

    def test_if_file_is_passed_it_is_returned_with_content_added(self):
        given_content = StringIO()
        assert given_content.getvalue() == ''
        self.mixin.generate_xlsx(self.data, file=given_content)
        assert given_content.getvalue() > ''

    def test_adds_row_of_data(self):
        wb = self.mixin.generate_xlsx(self.data)
        ws = self._get_sheet(wb)
        assert ws.cell('A1').value == 'row1col1'
        assert ws.cell('B2').value == 'row2col2'

    def test_inserts_headers_if_provided(self):
        headers = ('ColA', 'ColB')
        wb = self.mixin.generate_xlsx(self.data, headers)
        ws = self._get_sheet(wb)
        assert ws.cell('A1').value == 'ColA'
        assert ws.cell('B2').value == 'row1col2'


class GenerateCsvTests(TestCase):
    def setUp(self):
        self.data = (('row1col1', 'row1col2'), ('row2col1', 'row2col2'))
        self.mixin = SpreadsheetResponseMixin()

    def test_if_no_file_passed_in_stringio_is_returned(self):
        generated_csv = self.mixin.generate_csv(self.data)
        assert type(generated_csv) == type(StringIO())

    def test_if_file_is_passed_it_is_returned_with_content_added(self):
        given_content = mock.MagicMock()
        returned_file = self.mixin.generate_csv(self.data, file=given_content)
        assert returned_file == given_content

    def test_adds_row_of_data(self):
        expected_string = \
            'row1col1,row1col2\r\nrow2col1,row2col2\r\n'
        generated_csv = self.mixin.generate_csv(self.data)
        assert generated_csv.getvalue() == expected_string

    def test_inserts_headers_if_provided(self):
        headers = ('ColA', 'ColB')
        expected_string = \
            'ColA,ColB\r\nrow1col1,row1col2\r\nrow2col1,row2col2\r\n'
        generated_csv = self.mixin.generate_csv(self.data, headers)
        assert generated_csv.getvalue() == expected_string

    def test_handles_data_with_commas_correctly_by_putting_quotes(self):
        data = (('Title is, boo', '2'), )
        expected_string = \
            '"Title is, boo",2\r\n'
        generated_csv = self.mixin.generate_csv(data)
        assert generated_csv.getvalue() == expected_string

    def test_handles_data_with_quotes_correctly(self):
        data = (('Title is, "boo"', '2'), )
        expected_string = \
            '"Title is, ""boo""",2\r\n'
        generated_csv = self.mixin.generate_csv(data)
        assert generated_csv.getvalue() == expected_string

    def test_handles_data_in_unicode_correctly(self):
        data = ((u'Bumblebee is čmrlj', '2'), )
        expected_string = u'Bumblebee is čmrlj,2\r\n'.encode('utf-8')
        generated_csv = self.mixin.generate_csv(data)
        assert generated_csv.getvalue() == expected_string


class RenderSetupTests(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()
        self.mixin.generate_headers = mock.MagicMock()
        self.mixin.queryset = MockModel.objects.all()
        self.fields = (u'title',)

    def test_if_no_self_queryset_raise_improperlyconfigured(self):
        delattr(self.mixin, 'queryset')
        with pytest.raises(NotImplementedError):
            list(self.mixin.render_setup())

    def test_get_fields_is_called_with_kwargs(self):
        self.mixin.get_fields = mock.MagicMock()
        self.mixin.render_setup(a=1, b=2)
        self.mixin.get_fields.assert_called_once_with(a=1, b=2)

    def test_generate_data_is_called_once_with_fields_and_queryset(self):
        self.mixin.generate_data = mock.MagicMock()
        qs = MockModel.objects.all()
        self.mixin.render_excel_response(queryset=qs, fields=self.fields)
        self.mixin.generate_data.assert_called_once_with(fields=self.fields)

    def test_if_no_headers_passed_generate_headers_called(self):
        self.mixin.render_excel_response(fields=self.fields)
        self.mixin.generate_headers.assert_called_once_with(MockModel,
                                                            fields=self.fields)

    def test_returns_attachment_with_correct_filename(self):
        expected_disposition = 'attachment; filename="export.csv"'
        response = self.mixin.render_csv_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

        self.mixin.filename = 'data.dump'

        expected_disposition = 'attachment; filename="data.dump"'
        response = self.mixin.render_csv_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

        expected_disposition = 'attachment; filename="data.dump"'
        response = self.mixin.render_excel_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

        delattr(self.mixin, 'filename')
        self.mixin.filename_base = 'data.dump'

        expected_disposition = 'attachment; filename="data.dump.csv"'
        response = self.mixin.render_csv_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

        expected_disposition = 'attachment; filename="data.dump.xlsx"'
        response = self.mixin.render_excel_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition


class RenderExcelResponseTests(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()
        MockModelFactory()
        self.queryset = MockModel.objects.all()
        self.mixin.queryset = self.queryset

    def test_returns_httpresponse(self):
        assert type(self.mixin.render_excel_response()) == HttpResponse

    def test_returns_xlsx_content_type(self):
        expected_content_type = \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = self.mixin.render_excel_response()
        actual_content_type = response._headers['content-type'][1]
        assert actual_content_type == expected_content_type

    def test_returns_attachment_content_disposition(self):
        expected_disposition = 'attachment; filename="export.xlsx"'
        response = self.mixin.render_excel_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

    def test_get_filename_called_with_csv_parameter(self):
        self.mixin.get_filename = mock.MagicMock()
        self.mixin.render_excel_response()
        self.mixin.get_filename.assert_called_once_with(extension='xlsx')

    def test_generate_xslx_is_called_with_data(self):
        self.mixin.generate_xlsx = mock.MagicMock()
        self.mixin.render_excel_response()
        data = self.mixin.generate_data()
        mut = self.mixin.generate_xlsx
        assert mut.call_count == 1
        assert list(mut.call_args.__getnewargs__()[0][1]['data']) == list(data)

    def test_generate_xslx_is_called_with_headers(self):
        self.mixin.generate_xlsx = mock.MagicMock()
        headers = ('ColA', 'ColB')
        self.mixin.render_excel_response(headers=headers)
        mut = self.mixin.generate_xlsx
        assert mut.call_count == 1
        assert mut.call_args.__getnewargs__()[0][1]['headers'] == headers

    def test_generate_xslx_is_called_with_response(self):
        self.mixin.generate_xlsx = mock.MagicMock()
        self.mixin.render_excel_response()
        mut = self.mixin.generate_xlsx
        assert mut.call_count == 1
        assert type(mut.call_args.__getnewargs__()[0][1]['file']) \
            == HttpResponse


class RenderCsvResponseTests(TestCase):
    def setUp(self):
        self.author = MockAuthorFactory()
        MockModelFactory(author=self.author)
        self.queryset = MockModel.objects.all()
        self.mixin = SpreadsheetResponseMixin()
        self.mixin.queryset = self.queryset

    def test_returns_httpresponse(self):
        assert type(self.mixin.render_csv_response()) == HttpResponse

    def test_returns_csv_content_type(self):
        expected_content_type = 'text/csv'
        response = self.mixin.render_csv_response()
        actual_content_type = response._headers['content-type'][1]
        assert actual_content_type == expected_content_type

    def test_returns_attachment_content_disposition(self):
        expected_disposition = 'attachment; filename="export.csv"'
        response = self.mixin.render_csv_response()
        actual_disposition = response._headers['content-disposition'][1]
        assert actual_disposition == expected_disposition

    def test_get_filename_called_with_csv_parameter(self):
        self.mixin.get_filename = mock.MagicMock()
        self.mixin.render_csv_response()
        self.mixin.get_filename.assert_called_once_with(extension='csv')

    def test_generate_csv_is_called_with_data(self):
        mut = self.mixin.generate_csv = mock.MagicMock()
        self.mixin.render_csv_response()
        data = self.mixin.generate_data()
        assert mut.call_count == 1
        assert list(mut.call_args.__getnewargs__()[0][1]['data']) == list(data)

    def test_generate_csv_is_called_with_headers(self):
        self.mixin.generate_csv = mock.MagicMock()
        headers = ('ColA', 'ColB')
        self.mixin.render_csv_response(headers=headers)
        mut = self.mixin.generate_csv
        assert mut.call_count == 1
        assert mut.call_args.__getnewargs__()[0][1]['headers'] == headers

    def test_generate_csv_is_called_with_response(self):
        self.mixin.generate_csv = mock.MagicMock()
        self.mixin.render_csv_response()
        mut = self.mixin.generate_csv
        assert mut.call_count == 1
        assert type(mut.call_args.__getnewargs__()[0][1]['file']) \
            == HttpResponse


class GenerateHeadersTests(TestCase):
    def setUp(self):
        MockModelFactory()
        self.mixin = SpreadsheetResponseMixin()
        self.mixin.queryset = MockModel.objects.all()
        self.data = self.mixin.generate_data()

    def test_generate_headers_gets_headers_from_model_name(self):
        fields = self.mixin.get_fields(model=MockModel)
        assert self.mixin.generate_headers(MockModel, fields) == (u'Id', u'Title', u'Author')

    def test_generate_headers_keeps_fields_order(self):
        fields = ('title', 'id')
        headers = self.mixin.generate_headers(MockModel, fields=fields)
        assert headers == (u'Title', u'Id')

    def test_generate_headers_only_returns_fields_if_fields_is_passed(self):
        fields = ('title',)
        assert self.mixin.generate_headers(MockModel,
                                           fields=fields) == (u'Title', )

    def test_generate_headers_follows_foreign_keys(self):
        fields = ('title', 'author__name')
        headers = self.mixin.generate_headers(MockModel, fields)
        assert headers == (u'Title', u'Author Name')

    def test_generate_headers_with_calculated_fields(self):
        fields = ('title', 'author__name', 'calculate_this')
        self.mixin.calculate_this = lambda values: 'whee %d' % values[0]
        headers = self.mixin.generate_headers(MockModel, fields)
        assert headers == (u'Title', u'Author Name', u'Calculate This')

    def test_generate_headers_with_calculated_fields_with_verbose_names(self):
        fields = ('title', 'author__name', 'calculate_this')
        self.mixin.calculate_this = lambda values: 'whee %d' % values[0]
        self.mixin.calculate_this.verbose_name = 'Whee!'
        headers = self.mixin.generate_headers(MockModel, fields)
        assert headers == (u'Title', u'Author Name', u'Whee!')


class GetFieldsTests(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()

    def test_if_fields_defined_on_view(self):
        fields = ('title', 'summary')
        self.mixin.fields = fields
        assert self.mixin.get_fields() == fields

    def test_get_fields_from_kwargs(self):
        fields = ('title', 'summary')
        assert self.mixin.get_fields(fields=fields) == fields

    def test_get_fields_from_queryset(self):
        self.mixin.queryset = MockModel.objects.all()
        assert self.mixin.get_fields() == ['id', 'title', 'author']

    def test_get_fields_from_values_list_queryset(self):
        self.mixin.queryset = MockModel.objects.all().values_list()
        assert self.mixin.get_fields() == ['id', 'title', 'author_id']

    def test_get_fields_from_values_list_queryset_with_specific_fields(self):
        self.mixin.queryset = MockModel.objects.all().values_list('author', 'title')
        assert self.mixin.get_fields() == ['author', 'title']


class GetRenderMethodTest(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()

    def test_raise_notimplemented_for_unknown_format(self):
        with pytest.raises(NotImplementedError):
            self.mixin.get_render_method('doc')

    def test_returns_excel_response_method_for_excel_format(self):
        expected_render_method = self.mixin.render_excel_response
        assert self.mixin.get_render_method('excel') == expected_render_method

    def test_csv_response_method_for_csv_format(self):
        expected_render_method = self.mixin.render_csv_response
        assert self.mixin.get_render_method('csv') == expected_render_method


class GetFormatTest(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()

    def test_get_format_from_export_format_kwarg(self):
        format = 'excel'
        assert self.mixin.get_format(format=format) == format

    def test_get_format_from_export_format_attribute(self):
        format = 'csv'
        self.mixin.format = format
        assert self.mixin.get_format() == format

    def test_raise_notimplemented_if_export_format_not_supplied(self):
        with pytest.raises(NotImplementedError):
            self.mixin.get_format()


class GetFilenameTest(TestCase):
    def setUp(self):
        self.mixin = SpreadsheetResponseMixin()

    def test_from_filename_kwarg(self):
        filename = 'filename.kwarg'
        assert self.mixin.get_filename(filename=filename) == filename

    def test_from_filename_attribute(self):
        filename = 'filename.attr'
        self.mixin.filename = filename
        assert self.mixin.get_filename() == filename

    def test_default_return_with_no_extension_provided(self):
        assert self.mixin.get_filename() == 'export.out'

    def test_default_return_with_extension_provided(self):
        assert self.mixin.get_filename(extension='blob') == 'export.blob'
